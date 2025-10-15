from datetime import datetime, date
import shutil
import pandas as pd
import glob
import os
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from unidecode import unidecode
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# folder_path = r'C:\Temp\Cadastro de Tabela'

def converter_data(valor):
    try:
        if pd.isnull(valor):
            return pd.NaT
        if isinstance(valor, (int, float)):
            # Excel's date origin is '1899-12-30' for Windows, '1904-01-01' for Mac.
            # Assuming Windows origin.
            return pd.to_datetime(valor, origin='1899-12-30', unit='D')
        else:
            # Try to convert string to datetime, handling various formats
            data = pd.to_datetime(valor, dayfirst=True, errors='coerce')
            return data
    except Exception as e:
        print(f"Erro ao converter o valor: {valor}. Erro: {e}")
        return pd.NaT

# Defina a opção do pandas para aceitar o comportamento futuro
pd.set_option('future.no_silent_downcasting', True)

# Função para processar arquivo Excel
def processar_arquivo_excel(excel_path):
    # Definir a data atual no início da função para evitar o erro UnboundLocalError
    data_atual = pd.Timestamp(date.today())

    try:
        df = pd.read_excel(excel_path)
    except Exception as e:
        print(f"Erro ao ler o arquivo: {e}")
        return None, None
    
    col_idx = {
        'COD_TRANSPORTADORA': 0,
        'NOME_TRANSPORTADORA': 1,
        'TIPO_DE_TARIFA': 2,
        'TIPO_DE_DEDICADA': 3,
        '%_DO_TRAJETO_DA_FD': 4,
        'OPERACAO': 5,
        'ID_ORIGEM': 6,
        'ZONA_DE_TRANSPORTE_ORIGEM': 7,
        'DESCRICAO_ZONA_DE_TRANSPORTE_ORIGEM': 8,
        'ID_DESTINO': 9,
        'ZONA_DE_TRANSPORTE_DESTINO': 10,
        'DESCRICAO_ZONA_DE_TRANSPORTE_DESTINO': 11,
        'PERFIL_GRUPO_DE_EQUIPAMENTO': 12,
        'DESCRICAO_GRUPO_DE_EQUIPAMENTO': 13,
        'CUSTO_BASE': 14,
        'ABASTECIMENTO_CDS': 15,
        'REDESPACHO': 16,
        'REVISTA': 17,
        'AJUDANTE': 18,
        'PEDAGIO': 19,
        'PORCENTAGEM_AD_VALOREM': 20,
        'VALOR_DE_CONTRATO': 21,
        'NRO_DE_VEICULOS': 22,
        'TARIFA_PREFERENCIAL': 23,
        'DATA_DE_INICIO': 24,
        'DATA_DE_EXPIRACAO': 25,
        'NRO_COLETAS': 26,
        'NRO_ENTREGAS': 27
    }

   
    # Extrair colunas esperadas do dicionário col_idx
    colunas_esperadas = list(col_idx.keys())

   # Verificar colunas esperadas e mostrar aviso para cada coluna ausente
    colunas_faltantes = []
    for coluna in colunas_esperadas:
        if coluna not in df.columns:
            colunas_faltantes.append(coluna)
    if colunas_faltantes:
        print(f"Aviso: As seguintes colunas esperadas estão ausentes: {', '.join(colunas_faltantes)}")
        return None, None  # Interromper o processamento se colunas críticas estiverem ausentes


    # Identificar colunas inesperadas
    colunas_inesperadas = [col for col in df.columns if col not in colunas_esperadas]
    if colunas_inesperadas:
        print(f"Aviso: Existem colunas inesperadas: {', '.join(colunas_inesperadas)}")

    # Manter e tratar só colunas esperadas
    df = df[[col for col in df.columns if col in colunas_esperadas]]
    for coluna in colunas_esperadas:
        if coluna not in df.columns:
            df[coluna] = np.nan
   
    
    # Colunas de descrição com acentos a serem removidos
    colunas_com_descricoes = [
        'ID_ORIGEM',
        'ID_DESTINO',
        'DESCRICAO_ZONA_DE_TRANSPORTE_ORIGEM',
        'DESCRICAO_ZONA_DE_TRANSPORTE_DESTINO',
        'DESCRICAO_GRUPO_DE_EQUIPAMENTO'
    ]

    # Tratar as colunas ID_ORIGEM e ID_DESTINO
    for coluna in ['ID_ORIGEM', 'ID_DESTINO', 'COD_TRANSPORTADORA']:
        if coluna in df.columns:
            # df[coluna] = df[coluna].astype(str).apply(unidecode)
            df[coluna] = df[coluna].apply(
                lambda x: unidecode(str(int(x))) if isinstance(x, float) and x.is_integer() else unidecode(str(x))
            )

    # Remover acentos das demais colunas
    for coluna in colunas_com_descricoes[2:]:  # Começa do índice 2
        if coluna in df.columns:
            df[coluna] = df[coluna].apply(lambda x: unidecode(x) if isinstance(x, str) else x)
    
           
     # Limpar as colunas de strings e valores de 'N'
    df.replace('N', np.nan, inplace=True)
    
  
     # Garantir que a coluna 'CUSTO_BASE / CUSTO_VARIAVEL' esteja presente antes de operar
    custo_base = col_idx['CUSTO_BASE']
    if custo_base in df.columns:
        df.iloc[:, custo_base] = df.iloc[:, custo_base].apply(lambda x: str(x).replace(" ", "").replace('.', '').replace(',', '.') if isinstance(x, str) else x)
        df.iloc[:, custo_base] = pd.to_numeric(df.iloc[:, custo_base], errors='coerce').fillna(0).round(2)


    # Função para corrigir e converter colunas
    def corrigir_converter_coluna(df, coluna, tipo='numerico'):
        if coluna in df.columns:
            if tipo == 'numerico':
                df[coluna] = df[coluna].apply(lambda x: str(x).replace(" ", "").replace('.', '').replace(',', '.') if isinstance(x, str) else x)
                df[coluna] = pd.to_numeric(df[coluna], errors='coerce').fillna(0).round(2)
            elif tipo == 'espacos':
                df[coluna] = df[coluna].apply(lambda x: str(x).strip() if isinstance(x, str) else x)
        return df

    # Listas de colunas por tipo de tratamento
    colunas_numericas = ['ABASTECIMENTO_CDS', 'REDESPACHO']
    # colunas_espacos = ['COD_TRANSPORTADORA', 'OPERACAO', 'ZONA_DE_TRANSPORTE_ORIGEM', 'ZONA_DE_TRANSPORTE_DESTINO']

    # Aplicar função nas colunas numéricas
    for col in colunas_numericas:
        df = corrigir_converter_coluna(df, col, tipo='numerico')

    # Aplicar função nas colunas com remoção de espaços
    for col in df.columns:
        df = corrigir_converter_coluna(df, col, tipo='espacos')
    

    
    colunas_para_imprimir = {
        'COD_TRANSPORTADORA': 'TRP',
        'TIPO_DE_TARIFA': 'TARIFA',
        'DESCRICAO_ZONA_DE_TRANSPORTE_ORIGEM': 'ORIGEM',
        'DESCRICAO_ZONA_DE_TRANSPORTE_DESTINO': 'DESTINO',
        'DESCRICAO_GRUPO_DE_EQUIPAMENTO': 'VEÍCULO',
        'CUSTO_BASE': 'R$ FRETE',
        'ABASTECIMENTO_CDS': 'R$ PARADA',
        'REDESPACHO': 'R$ TROCA NF',
        'DATA_DE_INICIO': "INICIO",
        'DATA_DE_EXPIRACAO': "TERMINO",
        'NRO_COLETAS': 'COLETAS',
        'NRO_ENTREGAS': 'ENTREGAS'
    }

    # Verificar valores menores que 1 ou vazios
    valores_invalidos = df.iloc[:, custo_base] < 1
   
    # Se houver valores inválidos, imprimir mensagem de erro
    if valores_invalidos.any():
        print('------------------------------------------------------------------------------------')
        print("Erro: Arquivo possui valor de frete zerado. Seguem os registros com erro para correção e reprocesso!")
        print()
        # Filtrar apenas as linhas problemáticas para imprimir com colunas renomeadas
        df_valores_invalidos = df[valores_invalidos][list(colunas_para_imprimir.keys())].rename(columns=colunas_para_imprimir)
        print(df_valores_invalidos)
        exit() # Interrompe se houver valores de frete zerados
    else:
        # Se não houver valores inválidos, mostrar as colunas definidas
        print('------------------------------------------------------------------------------------')
        df_visualizado = df[list(colunas_para_imprimir.keys())].rename(columns=colunas_para_imprimir)
        print(df_visualizado.head())
        
    # Transformar as colunas 'NRO_COLETAS' e 'NRO_ENTREGAS' em números inteiros
    df[['NRO_COLETAS', 'NRO_ENTREGAS']] = df[['NRO_COLETAS', 'NRO_ENTREGAS']].apply(pd.to_numeric, errors='coerce').astype('Int64')
        # Adicionando a conversão de datas usando a função converter_data
    df['DATA_DE_INICIO'] = df['DATA_DE_INICIO'].apply(converter_data)
    df['DATA_DE_EXPIRACAO'] = df['DATA_DE_EXPIRACAO'].apply(converter_data)

    # Impressão de linhas com datas inválidas
    datas_invalidas = df[df['DATA_DE_INICIO'].isna() | df['DATA_DE_EXPIRACAO'].isna()]
    if not datas_invalidas.empty:
        print("Linhas com datas inválidas:")
        print(datas_invalidas[['DATA_DE_INICIO', 'DATA_DE_EXPIRACAO']])

    # Validação de datas
    if df['DATA_DE_INICIO'].isnull().all() or df['DATA_DE_EXPIRACAO'].isnull().all():
        print("Erro: As colunas de data estão vazias ou contêm apenas valores inválidos.")
        return None, None
    
    # # Verificar se alguma data é anterior à data_atual
    # if (df['DATA_DE_INICIO'] < data_atual).any() or (df['DATA_DE_EXPIRACAO'] < data_atual).any():
    #     print("Erro: O arquivo possui datas em DATA_DE_INICIO ou DATA_DE_EXPIRACAO anteriores à data atual.")
    #     return None, None   

    # Remover linhas duplicadas
    df = df.drop_duplicates()
    # Preencher valores ausentes com strings vazias
    df = df.fillna('')
    # Use infer_objects para ajustar os tipos
    df = df.infer_objects(copy=False)
    # Limpar espaços extras dos nomes das colunas
    df.columns = df.columns.str.strip()

    # Verificar a condição adicional para 'TIPO_DE_TARIFA' e colunas AI e AJ
    cond = (df['TIPO_DE_TARIFA'] == 'CONSOLIDADO') & (df['NRO_COLETAS'] < 2) & (df['NRO_ENTREGAS'] < 2)
    # Verifica se a condição foi satisfeita em alguma linha
    if cond.any():  # Condição em qualquer linha que atenda ao critério
        resposta = input("Números de coletas e entregas não estão consolidadas. Deseja continuar? (s/n): ")
        if resposta.lower() != 's':
            print("Processo interrompido.")
            return None, None
    
    # Verificar condições de custo de abastecimento
    cond_custo_abastecimento = (
        ((df['NRO_COLETAS'] > 1) & (df['ABASTECIMENTO_CDS'] <= 0)) |
        ((df['NRO_ENTREGAS'] > 1) & (df['ABASTECIMENTO_CDS'] <= 0))
    )
    
    if cond_custo_abastecimento.any():
        print('----------------------------------------------------------------------------------------------------------------------------------')
        print("\nErro: 'Tarifa CONSOLIDADA deve ter custo de Abastecimento. Necessário correção antes de continuar.")
        print("Seguem os registros com problemas:")
        # Filtrar e exibir as linhas problemáticas
        df_custo_abastecimento_invalido = df[cond_custo_abastecimento][[
            'NRO_COLETAS', 'NRO_ENTREGAS', 'ABASTECIMENTO_CDS'
        ]]
        print(df_custo_abastecimento_invalido)
        print('----------------------------------------------------------------------')
        return None, None

    # Mostrar a contagem de linhas do DataFrame processado
    print(f"Quantidade de linhas no Arquivo: {len(df)}")
    print('------------------------------------------------------------------------------------')
    registros = []
    for index, row in df.iterrows():
        registro = {
            'Origem': row['ZONA_DE_TRANSPORTE_ORIGEM'],
            'Destino': row['ZONA_DE_TRANSPORTE_DESTINO'],
            'Transportadora': row['COD_TRANSPORTADORA'],
            'Equipamento': row['PERFIL_GRUPO_DE_EQUIPAMENTO']
        }
        registros.append(registro)

    return df, registros


# Função para criar e salvar novo arquivo Excel
def salvar_arquivo_excel(df):
    
    if df is None:
        print("DataFrame está vazio. Não é possível salvar o arquivo.")
        return
    
    # Criar um novo Workbook do openpyxl
    wb = Workbook()
    ws = wb.active
    
    # Adicionar o DataFrame ao worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Obter índice das colunas pelo cabeçalho
    col_DATA_DE_INICIO = None
    col_DATA_DE_EXPIRACAO = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == "DATA_DE_INICIO":
            col_DATA_DE_INICIO = col[0].column
        if col[0].value == "DATA_DE_EXPIRACAO":
            col_DATA_DE_EXPIRACAO = col[0].column

    # Verificar se as colunas foram encontradas
    if col_DATA_DE_INICIO and col_DATA_DE_EXPIRACAO:
        for row in ws.iter_rows(min_row=2, min_col=col_DATA_DE_INICIO, max_col=col_DATA_DE_EXPIRACAO):
            for cell in row:
                if isinstance(cell.value, (date, datetime)):  # Verificar se o valor é uma data
                    cell.number_format = 'DD/MM/YYYY'  # Aplicar formato de data
    else:
        print("Datas não encontradas.")

    print("\nConcluindo Tratamento...")

    # Salvar o Workbook em um novo arquivo
    # output_path = os.path.join(folder_path, 

def mover_arquivos_para_historico(folder_path):
    # Caminho para a pasta "historico"
    historico_folder = os.path.join(folder_path, 'historico')
    # Criar pasta histórico se não existir
    if not os.path.exists(historico_folder):
        os.makedirs(historico_folder)

    # Obter todos os arquivos .xlsx no diretório especificado
    files = glob.glob(os.path.join(folder_path, '*.xlsx'))
    # Mover cada arquivo encontrado para a pasta "historico" com um timestamp
    for file_path in files:
        # Extrair o nome do arquivo
        file_name = os.path.basename(file_path)
        # Adicionar um timestamp ao nome do arquivo para torná-lo único
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_file_name = f"{os.path.splitext(file_name)[0]}_{timestamp}.xlsx"
        # Caminho completo do novo arquivo na pasta "historico"
        historico_path = os.path.join(historico_folder, new_file_name)
        try:
            shutil.move(file_path, historico_path)
            print(f"Arquivo {file_name} movido para: {historico_path}")
        except Exception as e:
            print(f"Erro ao mover arquivo {file_name}: {e}")

def selecionar_arquivo():
    root = Tk()
    root.withdraw()
    
    arquivo_selecionado = askopenfilename(
        title="Escolha o arquivo Excel",
        filetypes=[("Excel files", "*.xlsx *.xls")],
        initialdir=r'C:\Temp\Cadastro de Tabela'
    )
    
    root.destroy()
    
    if not arquivo_selecionado:
        print("Nenhum arquivo foi selecionado.")
        return None
    
    print(f"Arquivo selecionado: {arquivo_selecionado}")
    return arquivo_selecionado

# Função principal para processar os arquivos
def main():
    print("Iniciando o processamento de arquivos Excel...")

    caminho_arquivo_excel = selecionar_arquivo()  # Função que retorna o caminho do arquivo
    df, registros = processar_arquivo_excel(caminho_arquivo_excel)
    
    if not caminho_arquivo_excel:
        print("Operação cancelada. Nenhum arquivo foi selecionado.")
        return

    df, registros = processar_arquivo_excel(caminho_arquivo_excel)

    if df is not None and registros is not None:
        # Salvar o arquivo processado
        salvar_arquivo_excel(df, caminho_arquivo_excel)

        # Defina o folder_path usando o diretório do arquivo selecionado
        folder_path = os.path.dirname(caminho_arquivo_excel)
        # Mover arquivos originais para a pasta histórico
        mover_arquivos_para_historico(folder_path)
        print("\nProcessamento concluído com sucesso!")
        print()
    else:
        print("Não foi possível completar o processamento devido a erros.")

# Executar o programa
if __name__ == "__main__":
    main()
